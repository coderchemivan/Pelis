from logging import raiseExceptions
from bs4 import BeautifulSoup
import requests
import re
import openpyxl
import json

from scrapy import Selector


class Pelis():
    def lista_pelis(self,archivo):
        wb = openpyxl.load_workbook(archivo)
        hoja = wb.active
        lista = {}
        for i in range(4,hoja.max_row+1):
            lista[hoja.cell(row=i, column=1).value] = hoja.cell(row=i, column=12).value
        return lista


    def peliScraper(self,movie,indice_,año="",director=""):
        info_peli = {}
        r =  requests.get('https://www.google.com/search?q={} {} {} imdb'.format(movie,año,director))
        soup = BeautifulSoup(r.text,'lxml')
        
        lista_tags_a_google = soup.find_all('a')[0:30]
        

        link_movie = ''
        for tag_a in lista_tags_a_google:
            link_ = re.findall('(https://www.imdb.com/title/.+/)',tag_a.get('href'))
            if len(link_)==1:
                link_movie = link_
                break
        r = requests.get(link_movie[0])
        soup = BeautifulSoup(r.text,'lxml')
        #print(link_movie[0],movie,año)


        '''Titulo'''
        try:
            titulo = soup.find_all('div',attrs={'class':'sc-dae4a1bc-0 gwBsXc'})[0].text
            titulo = titulo.replace('\n','').replace('Original title:','').strip()
            info_peli['title'] = titulo
        except:
            info_peli['title'] = movie

        '''indice'''
        info_peli['id'] = indice_


        '''género'''
        generos_ = soup.find_all('span',attrs={'class':'ipc-chip__text'})
        generos = []
        for genero in generos_:
            generos.append(genero.text)
        info_peli['genre'] = generos

        
        '''dirección'''
        directores= soup.find_all('ul',
            attrs={'class':'ipc-inline-list ipc-inline-list--show-dividers ipc-inline-list--inline ipc-metadata-list-item__list-content base'})[1]
        direccion = []
        directores = directores.find_all('li')
        for director in directores:
            direccion.append(director.text)
        info_peli['director'] = direccion
        

        '''año'''
        año = soup.find_all('span',attrs={'class':'sc-8c396aa2-2 itZqyK'})[0].text
        info_peli['year'] = año
        
        
        '''puntuación IMDB'''
        puntuacion = soup.find_all('span',attrs={'class':'sc-7ab21ed2-1 jGRxWM'})[0].text
        info_peli['rating'] = puntuacion


        '''taquilla'''
        try:
            budget = soup.find_all('li',attrs={'class','ipc-metadata-list__item sc-6d4f3f8c-2 fJEELB'})[0]
            budget = budget.div.ul.li.span.text
            info_peli['budget'] = budget

            boxOffice_collection = soup.find_all('li',attrs={'class','ipc-metadata-list__item sc-6d4f3f8c-2 fJEELB'})[1]
            boxOffice_collection = boxOffice_collection.div.ul.li.span.text
            info_peli['boxOffice_collection'] = boxOffice_collection
        except:
            pass
            
        '''sinopsis'''
        sinopsis = soup.find_all('div',attrs={'class','ipc-html-content-inner-div'})[0].text
        info_peli['plot'] = sinopsis
        

        '''link cast principal'''
        link_cast = 'https://www.imdb.com' + soup.find_all('div',
            attrs={'class':'ipc-title__wrapper'})[1].find('a').get('href')
            


        '''Descargando portada'''

        previewImageElements = soup.find_all('div',
            attrs={'class':'ipc-poster ipc-poster--baseAlt ipc-poster--dynamic-width sc-d383958-0 gvOdLN celwidget ipc-sub-grid-item ipc-sub-grid-item--span-2'})[0].find('a').get('href')
        
    
        r = requests.get('https://www.imdb.com' + previewImageElements)
        soup = BeautifulSoup(r.text,'lxml')
        try:
            portada = soup.find('div',
                attrs={'class':'sc-7c0a9e7c-2 bkptFa'}).find('img').get('src')
        except:
            pass
            

        r = requests.get(portada)
        if r.status_code == 200:
            with open('Application/files/portadas/{}_{}.jpg'.format(movie,año),'wb') as f:
                f.write(r.content)
                downloaded = True


        '''Actores principales'''
        r = requests.get(link_cast)
        soup = BeautifulSoup(r.text,'lxml')
        actores = soup.find_all('table',attrs={'class':'cast_list'})[0].find_all('tr')
        actores = actores[1:11]
        actores_principales = []
        for actor in actores:
            try:
                actores_principales.append(actor.find_all('td')[1].text.strip())
            except:
                continue
        info_peli['cast'] = actores_principales
        return info_peli
    def dbPelis(self):
        pass


peli = Pelis()
lista = peli.lista_pelis(r'C:\Users\ivan_\OneDrive - UNIVERSIDAD NACIONAL AUTÓNOMA DE MÉXICO\Desktop\pel.xlsm')
c = 0
coleccion_pelis = []
for pelicula,año in lista.items():
    if año is None:
        año = ''
    try:
        peli_info = peli.peliScraper(str(pelicula).strip(),c+1,str(año))
        print(peli_info)
        print('')
        coleccion_pelis.append(peli_info)
        c+=1
    except:
         print('No se pudo descargar la info de {}'.format(pelicula))
         continue

with open('Application/files/mis_peliculas.json', 'w') as fp:
    json.dump(coleccion_pelis, fp)

#peli.peliScraper('Memento',1,'2000')







